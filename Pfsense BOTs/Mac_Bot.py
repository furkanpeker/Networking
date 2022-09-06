
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from openpyxl import load_workbook
import time

# Global declerations for reading from excel
descDoc = ''
pc_macDoc = ''
ph_macDoc = ''
exp_dateDoc = ''
people = input("Enter number of people: ")

# Global declerations for writing to freeRadius
PATH="C:\Program Files\chromedriver.exe"
net_addr = input("Enter the ip addres of web authenticator site : ")
uname = input("Enter Pfsense username : ")
password = input("Enter Pfsense password : ")



driver = webdriver.Chrome(PATH)
# The portion of signing in to system
driver.get("https://"+ net_addr +"/")
buttonAdvanced = driver.find_element_by_xpath('//*[@id="details-button"]')
buttonAdvanced.click()
buttonProceed = driver.find_element_by_xpath('//*[@id="proceed-link"]')
buttonProceed.click()

# Sign in process;
uName = driver.find_element_by_xpath('//*[@id="usernamefld"]')
uName.send_keys(uname)

passwd = driver.find_element_by_xpath('//*[@id="passwordfld"]')
passwd.send_keys(password)
passwd.send_keys(Keys.RETURN)


linkServices = driver.find_element_by_xpath('//*[@id="pf-navbar"]/ul[1]/li[4]/a')
linkServices.click() # Services

linkRadius = driver.find_element_by_xpath('//*[@id="pf-navbar"]/ul[1]/li[4]/ul/li[10]/a')
linkRadius.click() # freeRadius

buttonMacs = driver.find_element_by_xpath('//*[@id="3"]/div/ul/li[2]/a')
buttonMacs.click() # MACs

wb=load_workbook("C:\\Users\\W_nitro\\Desktop\\Mac_bot\\Cihaz MAC Adresi Kayıt Formu.xlsx")
ws=wb.active

row = 1
while (row < 3):
    # The portion of reading data of a device from e-table
    pc_macDoc = ws["B" + str(row)].value
    descDoc = ws["A" + str(row)].value + " Pc"
    exp_dateDoc = ws["D" + str(row)].value
    # Clicking to ADD button to add a new device
    buttonAdd = driver.find_element_by_xpath('//*[@id="mainarea"]/tbody/tr/td[2]/table/tbody/tr/td/a')
    buttonAdd.click()
    print(pc_macDoc)
    if pc_macDoc == None:
        continue
    else:
        pc_macSite = driver.find_element_by_xpath('//*[@id="varmacsaddress"]')
        pc_macSite.send_keys(pc_macDoc)

        descSite = driver.find_element_by_xpath('//*[@id="description"]')
        descSite.send_keys(descDoc)

        exp_dateSite = driver.find_element_by_xpath('//*[@id="varmacsexpiration"]')
        exp_dateSite.send_keys(exp_dateDoc)

        # Then to save it clicks the save button
        buttonSave = driver.find_element_by_xpath('//*[@id="submit"]')
        buttonSave.click()
        # wb.save("C:\\Users\\W_nitro\\Desktop\\Mac_bot\\Cihaz MAC Adresi Kayıt Formu.xlsx")
    row+=1

row2 = 1
while (row2 < 3):
    # The portion of reading data of a device from e-table
    ph_macDoc = ws["C" + str(row)].value
    descDoc = ws["A" + str(row)].value + " Ph"
    exp_dateDoc = ws["D" + str(row)].value
    # Clicking to ADD button to add a new device
    buttonAdd = driver.find_element_by_xpath('//*[@id="mainarea"]/tbody/tr/td[2]/table/tbody/tr/td/a')
    buttonAdd.click()

    if ph_macDoc == None:
        continue
    else:
        ph_macSite = driver.find_element_by_xpath('//*[@id="varmacsaddress"]')
        ph_macSite.send_keys(ph_macDoc)

        descSite = driver.find_element_by_xpath('//*[@id="description"]')
        descSite.send_keys(descDoc)

        exp_dateSite = driver.find_element_by_xpath('//*[@id="varmacsexpiration"]')
        exp_dateSite.send_keys(exp_dateDoc)

        # Then to save it clicks the save button
        buttonSave = driver.find_element_by_xpath('//*[@id="submit"]')
        buttonSave.click()
    row2+=1
